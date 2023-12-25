from openpyxl import load_workbook

# 你的字符串列表
str_list = ['a', 'b', 'c', 'd']

# 加载已存在的xlsx文件
book = load_workbook('cs.xlsx')

# 获取sheet
sheet = book['Sheet1']

# 遍历字符串列表，一个个写入数据
for i, value in enumerate(str_list, start=1):
    # 写入数据到第二列，行数为i
    sheet.cell(row=i, column=2, value=value)
    # 保存xlsx文件
    book.save('cs.xlsx')

book.close()
