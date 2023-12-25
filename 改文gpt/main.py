import requests
import re
import json
import pandas as pd
from openpyxl import load_workbook

# 读取xlsx文件，跳过标头和列表序号
df = pd.read_excel('txt.xlsx', header=None, index_col=None)

headers = {
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Authorization': 'Bear',
    'Cache-Control': 'no-cache',
    'Connection': 'keep-alive',
    'Content-Type': 'application/json',
    'Origin': 'https://chat.daifuku.asia',
    'Pragma': 'no-cache',
    'Referer': 'https://chat.daifuku.asia/',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'cross-site',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36',
    'accept': 'text/event-stream',
    'sec-ch-ua': '"Chromium";v="118", "Google Chrome";v="118", "Not=A?Brand";v="99"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'x-requested-with': 'XMLHttpRequest',
}

with open('tishici.txt', 'r', encoding='utf-8') as f:
    tishici = f.read()

with open('setting.json', 'r', encoding='utf-8') as f:
    setting = json.load(f)

chongzhi_num = input('是否要重置zhongduan_jixu_num？（y/n）')
if chongzhi_num == 'y':
    setting['zhongduan_jixu_num'] = 1
    with open('setting.json', 'w', encoding='utf-8') as f:
        json.dump(setting, f, ensure_ascii=False, indent=4)

with open('setting.json', 'r', encoding='utf-8') as f:
    setting = json.load(f)

yuanju_lst = []
for index, row in df.items():
    for num, x in enumerate(row, start=1):
        if setting['zhongduan_jixu_num'] <= num:
            yuanju_lst.append(x)
    break

# 加载已存在的xlsx文件
book = load_workbook('txt.xlsx')

# 获取sheet
sheet = book['Sheet']

json_data = {
    'messages': [
        {
            'role': 'system',
            'content': tishici,
        }
    ],
    'stream': True,
    'model': 'gpt-3.5-turbo-0301',
    'temperature': 0.5,
    'presence_penalty': 0,
    'frequency_penalty': 0,
    'top_p': 1,
}
get_num = 2

jiequ_cishu = setting['zhongduan_jixu_num'] - 2
while jiequ_cishu > 0:
    json_data['messages'].append(
        {'role': 'user', 'content': "{'%s': '%s'}" % (jiequ_cishu, sheet.cell(row=jiequ_cishu, column=1).value)})
    json_data['messages'].append(
        {'role': 'assistant', 'content': "{'%s': '%s'}" % (jiequ_cishu, sheet.cell(row=jiequ_cishu, column=2).value)})
    jiequ_cishu += 1
    get_num -= 1
    if get_num == 0:
        break

for index, juzi in enumerate(yuanju_lst, start=setting['zhongduan_jixu_num']):
    print('输入：', "{'%s': '%s'}" % (str(index), juzi))
    user = {'role': 'user', 'content': "{'%s': '%s'}" % (str(index), juzi)}

    if len(json_data['messages']) > 7:
        del json_data['messages'][2]
        del json_data['messages'][1]

    json_data['messages'].append(user)
    chongxin_num = 5
    while chongxin_num:
        response = requests.post('https://cdn.daifuku.asia/v1/chat/completions', headers=headers, json=json_data)
        if response.status_code == 200:
            try:
                text_lst = re.findall(r'content":"(.*?)"},"finish', response.content.decode('utf-8'))
                print('返回：', ''.join(text_lst))
                assistant_json_lst = re.findall(r'{\'(.*?)\'}|{\"(.*?)\"}', ''.join(text_lst).replace('\\', ''))[0]
                if assistant_json_lst[0] != '':
                    assistant_content = "{'%s'}" % assistant_json_lst[0]
                else:
                    assistant_content = '{"%s"}' % assistant_json_lst[1]
                assistant_content = eval(assistant_content)[str(index)]
            except:
                print('')
                print('')
            assistant = {'role': 'assistant', 'content': "{'%s': '%s'}" % (str(index), assistant_content)}
            json_data['messages'].append(assistant)
            print('-------------------------')
            # 写入数据到第二列，行数为index
            sheet.cell(row=index, column=2, value=assistant_content)
            # 保存xlsx文件
            book.save('txt.xlsx')
            setting['zhongduan_jixu_num'] += 1
            with open('setting.json', 'w', encoding='utf-8') as f:
                json.dump(setting, f, ensure_ascii=False, indent=4)
            break
        else:
            chongxin_num -= 1
            print('不为200，次数还有：', chongxin_num)

    if not chongxin_num:  # 如果为0
        print('#####出现为0，终止#####')
        break

book.close()
